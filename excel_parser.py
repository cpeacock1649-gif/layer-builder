"""
Excel Parser for Insurance Program Structures

This module handles parsing of insurance program Excel files to extract:
- Insurance layers (limits and attachment points)
- Carrier information and participation
- Premium amounts and fees
- Shares and allocations

Enhanced to support OHSU-style broker program schedules with:
- Layer headers like "$75M ex $100M EQ", "$100M ex $300M", "$500M ex $1BL"
- Support for trailing text: EQ, AR ex, Terrorism, ALL RISKS, etc.
- Flexible column detection for Participant, Line, PPM, Premium, Fees, SL Tax, Total
"""

import re
from typing import Dict, List, Optional, Any, Tuple
from openpyxl import load_workbook
import io
import logging

# Debug flag - set to True to enable detailed logging
DEBUG_PARSING = False

# Configure logging
logger = logging.getLogger(__name__)
if DEBUG_PARSING:
    logging.basicConfig(level=logging.DEBUG)
else:
    logging.basicConfig(level=logging.WARNING)


def parse_currency(value) -> float:
    """
    Parse currency value from various formats, including Billions (B/BL).

    Args:
        value: Currency value (string, number, or None)

    Returns:
        Float value
    """
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    # Remove currency symbols, commas, and whitespace
    value_str = str(value).replace("$", "").replace(",", "").strip()

    # Handle suffixes - order matters: check BL before B, MM before M
    multiplier = 1
    upper_str = value_str.upper()

    if "MM" in upper_str:
        # MM = millions (alternative notation)
        multiplier = 1_000_000
        value_str = upper_str.replace("MM", "")
    elif "BL" in upper_str:
        # BL = billions
        multiplier = 1_000_000_000
        value_str = upper_str.replace("BL", "")
    elif "B" in upper_str and "BL" not in upper_str:
        # B = billions (but not if it's part of BL)
        multiplier = 1_000_000_000
        value_str = upper_str.replace("B", "")
    elif "M" in upper_str:
        multiplier = 1_000_000
        value_str = upper_str.replace("M", "")
    elif "K" in upper_str:
        multiplier = 1_000
        value_str = upper_str.replace("K", "")

    # Clean any remaining non-numeric characters except decimal point and minus
    value_str = re.sub(r"[^\d.\-]", "", value_str)

    try:
        return float(value_str) * multiplier if value_str else 0.0
    except (ValueError, AttributeError):
        return 0.0


def extract_layer_from_text(text: str) -> Optional[Dict[str, Any]]:
    """
    Extract layer information from text like "$75M ex $100M EQ" or "$500M ex $1BL".

    Enhanced to handle OHSU-style patterns including:
    - "$75M ex $100M EQ" - with trailing descriptors
    - "$100M ex $300M" - standard excess format
    - "$500M ex $1.5BL AR ex" - with AR (All Risks) excess notation
    - "$250M Terrorism" - standalone limit with description
    - "ALL RISKS EX ZURICH LEAD" - special layer types
    - "$1BL" - primary layer notation

    Args:
        text: Text containing layer information

    Returns:
        Dictionary with limit and attachment, or None
    """
    if not text or not isinstance(text, str):
        return None

    text = str(text).strip()

    if DEBUG_PARSING:
        logger.debug(f"Attempting to extract layer from: '{text}'")

    # Pattern 1: Standard excess format - $XM/K/B/BL ex/xs/excess $YM/K/B/BL (with optional trailing text like EQ, AR, etc.)
    # This handles: "$75M ex $100M EQ", "$100M ex $300M", "$500M ex $1.5BL AR ex"
    excess_patterns = [
        r"\$?([\d,]+(?:\.\d+)?)\s*([MKB](?:L)?|MM)?\s*(?:ex|excess|xs|excess\s+of)\s*\$?([\d,]+(?:\.\d+)?)\s*([MKB](?:L)?|MM)?",
    ]

    for pattern in excess_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            limit_num = match.group(1)
            limit_suffix = match.group(2) or ""
            attach_num = match.group(3)
            attach_suffix = match.group(4) or ""

            limit = parse_currency(f"{limit_num}{limit_suffix}")
            attachment = parse_currency(f"{attach_num}{attach_suffix}")

            if DEBUG_PARSING:
                logger.debug(
                    f"  Matched excess pattern: limit={limit}, attachment={attachment}"
                )

            return {
                "limit": limit,
                "attachment": attachment,
                "is_primary": False,
                "raw_text": text.strip(),
            }

    # Pattern 2: Primary layer patterns - "$XM primary", "$XM Primary Layer", etc.
    primary_patterns = [
        r"\$?([\d,]+(?:\.\d+)?)\s*([MKB](?:L)?|MM)?\s*(?:primary|primary\s+layer)",
    ]

    for pattern in primary_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            limit_num = match.group(1)
            limit_suffix = match.group(2) or ""
            limit = parse_currency(f"{limit_num}{limit_suffix}")

            if DEBUG_PARSING:
                logger.debug(f"  Matched primary pattern: limit={limit}")

            return {
                "limit": limit,
                "attachment": 0,
                "is_primary": True,
                "raw_text": text.strip(),
            }

    # Pattern 3: Standalone limit with description - "$250M Terrorism", "$1BL" at start
    # This handles special layers like terrorism coverage
    standalone_pattern = r"^\$?([\d,]+(?:\.\d+)?)\s*([MKB](?:L)?|MM)?\s*(?:terrorism|all\s*risk|property|liability|umbrella|excess)?"
    match = re.search(standalone_pattern, text, re.IGNORECASE)
    if match:
        limit_num = match.group(1)
        limit_suffix = match.group(2) or ""
        limit = parse_currency(f"{limit_num}{limit_suffix}")

        # Check if "primary" is mentioned or if this appears to be a primary layer
        is_primary = "primary" in text.lower() or (
            limit > 0 and "ex" not in text.lower() and "xs" not in text.lower()
        )

        # Only return if we have a valid limit and this looks like a layer definition
        if limit > 0:
            # Check for keywords that indicate this is a layer header
            layer_keywords = [
                "terrorism",
                "all risk",
                "property",
                "liability",
                "umbrella",
                "excess",
                "primary",
                "lead",
            ]
            if any(kw in text.lower() for kw in layer_keywords) or re.search(
                r"^\$[\d,]+(?:\.\d+)?[MKB]?(?:L)?$", text.strip(), re.IGNORECASE
            ):
                if DEBUG_PARSING:
                    logger.debug(
                        f"  Matched standalone pattern: limit={limit}, is_primary={is_primary}"
                    )
                return {
                    "limit": limit,
                    "attachment": 0,
                    "is_primary": is_primary,
                    "raw_text": text.strip(),
                }

    # Pattern 4: "ALL RISKS EX ZURICH LEAD" style - special named layers
    # These typically need the limit from a different cell/column
    all_risks_pattern = r"all\s*risks?\s*(?:ex|excess|xs)?\s*[\w\s]*(?:lead|primary)?"
    if re.search(all_risks_pattern, text, re.IGNORECASE):
        if DEBUG_PARSING:
            logger.debug(f"  Matched ALL RISKS pattern (no limit in text)")
        # Return None here - the layer info should be extracted from other cells
        # But mark this as detected for the caller
        return None

    if DEBUG_PARSING:
        logger.debug(f"  No layer pattern matched")

    return None


def is_layer_header_row(row_values: list) -> Tuple[bool, Optional[Dict[str, Any]]]:
    """
    Determine if a row contains a layer definition.

    Enhanced to handle OHSU-style headers including:
    - Standard excess layers: "$75M ex $100M EQ"
    - Named layers: "ALL RISKS EX ZURICH LEAD"
    - Layer headers in first column with limit/line data in subsequent columns

    Returns:
        Tuple of (is_layer_header: bool, layer_info: Optional[Dict])
    """
    if not row_values or not any(row_values):
        return False, None

    if DEBUG_PARSING:
        logger.debug(f"Checking if layer header row: {row_values[:6]}")

    # Only check the FIRST 2 cells for layer pattern - layer headers are typically
    # in the first column, not buried in the middle of data
    for i in range(min(2, len(row_values))):
        cell_value = row_values[i]
        if cell_value and isinstance(cell_value, str):
            cell_str = str(cell_value).strip()

            # Skip if this looks like a carrier name (contains "Insurance", "Company", etc.)
            carrier_indicators = [
                "insurance",
                "company",
                "assurance",
                "underwriter",
                "syndicate",
                "lloyds",
                "lloyd's",
                "inc",
                "ltd",
                "corp",
            ]
            cell_lower = cell_str.lower()
            if any(ind in cell_lower for ind in carrier_indicators):
                continue

            # Try to extract layer info from this cell
            layer_info = extract_layer_from_text(cell_str)
            if layer_info:
                if DEBUG_PARSING:
                    logger.debug(f"  Found layer in cell {i}: {layer_info}")
                return True, layer_info

            # Check for patterns like "75M ex 100M" without $ (common in some spreadsheets)
            if re.search(
                r"\d+[MKB]?(?:L)?\s*(?:ex|excess|xs)\s*\d+[MKB]?",
                cell_str,
                re.IGNORECASE,
            ):
                # Try to parse with added $ signs
                modified = re.sub(r"(\d+[MKB]?(?:L)?)", r"$\1", cell_str)
                layer_info = extract_layer_from_text(modified)
                if layer_info:
                    if DEBUG_PARSING:
                        logger.debug(
                            f"  Found layer (modified) in cell {i}: {layer_info}"
                        )
                    return True, layer_info

            # Check for special layer types that might not have amounts in this cell
            special_layer_patterns = [
                r"^all\s*risks?\s*(?:ex|excess|xs)",
                r"^\$?\d+[MKB]?\s*terrorism",
                r"^primary.*(?:all\s*risk|including|flood|eq)",
            ]
            for pattern in special_layer_patterns:
                if re.search(pattern, cell_str, re.IGNORECASE):
                    if DEBUG_PARSING:
                        logger.debug(
                            f"  Found special layer pattern in cell {i}: {pattern}"
                        )
                    # Try to find limit in other cells of this row
                    for j in range(len(row_values)):
                        if j != i and row_values[j]:
                            val = row_values[j]
                            if isinstance(val, (int, float)) and val >= 1_000_000:
                                # This might be the limit (must be at least $1M)
                                return True, {
                                    "limit": float(val),
                                    "attachment": 0,
                                    "is_primary": "primary" in cell_str.lower(),
                                    "raw_text": cell_str,
                                }
                            elif isinstance(val, str):
                                parsed = parse_currency(val)
                                if parsed >= 1_000_000:
                                    return True, {
                                        "limit": parsed,
                                        "attachment": 0,
                                        "is_primary": "primary" in cell_str.lower(),
                                        "raw_text": cell_str,
                                    }
                    # Return as layer header even without limit (will be handled later)
                    return True, None

    return False, None


def is_participant_header_row(row_values: list) -> bool:
    """
    Determine if a row is a participant/carrier header row.

    Enhanced to detect OHSU-style headers with columns like:
    - Participant, Line, PPM, Premium, Fees, SL tax, Total
    - Carrier, Share, Rate, etc.
    """
    if not row_values or not any(row_values):
        return False

    # Keywords that indicate a header row - must be exact cell matches or close
    # We check individual cells, not the whole row text, to avoid false positives
    header_keywords = {
        "participant",
        "line",
        "ppm",
        "premium",
        "fees",
        "fee",
        "carrier",
        "share",
        "firm",
        "sl tax",
        "surplus",
        "total",
        "rate",
    }

    # Count cells that exactly match header keywords
    keyword_matches = 0
    for cell in row_values:
        if cell:
            cell_text = str(cell).lower().strip()
            # Check for exact match or close match (e.g., "Line" or "Premium ($)")
            cell_clean = re.sub(r"[^a-z\s]", "", cell_text).strip()
            if cell_clean in header_keywords:
                keyword_matches += 1
            # Also check if cell starts with a keyword followed by space/punctuation
            for kw in header_keywords:
                if (
                    cell_clean == kw
                    or cell_text.startswith(kw + " ")
                    or cell_text.startswith(kw + "(")
                ):
                    keyword_matches += 1
                    break

    # Need at least 3 header keyword cells to consider this a header row
    # This is stricter to avoid false positives
    is_header = keyword_matches >= 3

    if DEBUG_PARSING and is_header:
        logger.debug(f"Detected participant header row: {row_values[:8]}")

    return is_header


def is_total_row(row_values: list) -> bool:
    """
    Determine if a row is a total/summary row by checking the first few cells.

    Enhanced to handle:
    - TOTAL, TOTALS, SUBTOTAL, SUM
    - GRAND TOTAL (for multi-layer summaries)
    """
    if not row_values:
        return False

    # Check first 4 cells for total indicators
    for i in range(min(4, len(row_values))):
        cell_val = str(row_values[i]).strip().upper() if row_values[i] else ""
        if cell_val in [
            "TOTAL",
            "TOTALS",
            "SUBTOTAL",
            "SUM",
            "GRAND TOTAL",
            "LAYER TOTAL",
        ]:
            if DEBUG_PARSING:
                logger.debug(f"Detected total row at cell {i}: {cell_val}")
            return True

    return False


def is_skip_row(row_values: list) -> bool:
    """
    Determine if a row should be skipped (empty, divider, or metadata row).
    """
    if not row_values or not any(row_values):
        return True

    # Check if row only contains formatting/metadata
    non_empty = [v for v in row_values if v is not None and str(v).strip()]
    if not non_empty:
        return True

    # Skip rows that are likely dividers or section headers without data
    first_val = str(row_values[0]).strip().lower() if row_values[0] else ""
    skip_patterns = [
        r"^[-=_\s]+$",  # Divider rows
        r"^note[s]?:",  # Note rows
        r"^comment[s]?:",
        r"^see\s+",  # Reference rows
    ]

    for pattern in skip_patterns:
        if re.match(pattern, first_val, re.IGNORECASE):
            return True

    return False


def map_columns(row_values: list) -> Dict[str, int]:
    """
    Map column headers to indices for data extraction.

    Enhanced to handle OHSU-style column variations:
    - Participant / Carrier / Firm
    - Line (participation amount)
    - PPM (price per million) / Rate
    - Premium
    - Fees / Fee
    - SL tax / SL Tax / Surplus Lines Tax
    - Total

    IMPORTANT: Column A (index 0) typically contains the layer header,
    so we skip it and only map columns B onwards (index 1+).
    """
    col_map = {}

    # Determine starting column - check if column 0 contains layer-like text
    start_col = 0
    if len(row_values) > 0 and row_values[0]:
        first_cell = str(row_values[0]).strip()
        first_cell_lower = first_cell.lower()

        # If first cell contains layer pattern like "$XM ex $XM", skip it
        if re.search(
            r"\$?\d+[mkb]?\s*(?:ex|xs|excess)", first_cell_lower, re.IGNORECASE
        ):
            start_col = 1
        # If first cell is NOT a header keyword but other cells look like headers, skip col 0
        elif first_cell_lower not in [
            "participant",
            "carrier",
            "firm",
            "line",
            "premium",
        ]:
            # Check if cells 1-4 contain header keywords
            for i in range(1, min(5, len(row_values))):
                if row_values[i]:
                    cell_text = str(row_values[i]).lower().strip()
                    if cell_text in ["participant", "line", "premium"]:
                        start_col = 1
                        break

    for col_idx, cell_value in enumerate(row_values):
        # Skip columns before start_col (typically column A with layer header)
        if col_idx < start_col:
            continue

        if not cell_value:
            continue

        cell_text = str(cell_value).lower().strip()
        # Clean cell text - remove non-alphabetic chars for exact matching
        cell_clean = re.sub(r"[^a-z\s]", "", cell_text).strip()

        # Carrier/Participant column - EXACT match only
        if cell_clean in ["participant", "participants", "carrier", "carriers", "firm"]:
            col_map["carrier"] = col_idx

        # Line column - EXACT match only
        elif cell_clean in ["line", "lines"]:
            col_map["line"] = col_idx

        # PPM (Price Per Million) / Rate column
        elif cell_clean in ["ppm", "rate"]:
            col_map["ppm"] = col_idx

        # Premium column - EXACT match only
        elif cell_clean in ["premium", "premiums"]:
            col_map["premium"] = col_idx

        # Fees column (but not SL fees) - EXACT match only
        elif cell_clean in ["fee", "fees"]:
            col_map["fees"] = col_idx

        # SL Tax / Surplus Lines Tax column
        elif cell_clean in ["sl tax", "sl taxes"]:
            col_map["sl_tax"] = col_idx
        elif "sl" in cell_text and "tax" in cell_text:
            col_map["sl_tax"] = col_idx

        # Total column - EXACT match only
        elif cell_clean in ["total", "totals"]:
            col_map["total"] = col_idx

    # Store the data start column so we know to skip column A when reading carrier data
    col_map["_data_start_col"] = start_col

    if DEBUG_PARSING:
        logger.debug(f"Column mapping (data starts at col {start_col}): {col_map}")

    return col_map


def parse_excel_program(
    excel_bytes: bytes, filename: str = "", debug: bool = False
) -> Dict[str, Any]:
    """
    Parse an insurance program Excel file.

    Enhanced to handle OHSU-style broker program schedules with:
    - Multiple layers per sheet
    - Layer headers with various formats
    - Participant rows with carrier data
    - Total rows that should be skipped
    - Support for multiple sheets

    Args:
        excel_bytes: Excel file as bytes
        filename: Original filename (optional)
        debug: Enable debug logging for this parse

    Returns:
        Dictionary containing extracted program data
    """
    global DEBUG_PARSING
    original_debug = DEBUG_PARSING
    if debug:
        DEBUG_PARSING = True
        logging.getLogger(__name__).setLevel(logging.DEBUG)

    try:
        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)

        all_layers = []

        # Process all sheets (not just the active one)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if DEBUG_PARSING:
                logger.debug(f"\n{'='*50}")
                logger.debug(f"Processing sheet: {sheet_name}")
                logger.debug(f"{'='*50}")

            current_layer = None
            current_layer_info = None
            col_map = {}

            # Iterate through rows
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(row):  # Skip empty rows
                    continue

                row_list = list(row)

                if DEBUG_PARSING:
                    logger.debug(f"\nRow {row_idx}: {row_list[:8]}")

                # 1. Check if this row should be skipped
                if is_skip_row(row_list):
                    if DEBUG_PARSING:
                        logger.debug(f"  -> Skipping (skip row)")
                    continue

                # 2. Check for Total Row first (to skip processing it as a carrier)
                if is_total_row(row_list):
                    if DEBUG_PARSING:
                        logger.debug(f"  -> Skipping (total row)")
                    continue

                # 3. Check if this is a layer header
                is_layer, layer_info = is_layer_header_row(row_list)
                if is_layer:
                    if DEBUG_PARSING:
                        logger.debug(f"  -> Layer header detected: {layer_info}")

                    # Save previous layer if it has carriers
                    if current_layer and current_layer.get("carriers"):
                        all_layers.append(current_layer)
                        if DEBUG_PARSING:
                            logger.debug(
                                f"  -> Saved previous layer with {len(current_layer['carriers'])} carriers"
                            )

                    if layer_info:
                        current_layer_info = layer_info
                        current_layer = {
                            "limit": layer_info["limit"],
                            "attachment": layer_info["attachment"],
                            "is_primary": layer_info["is_primary"],
                            "carriers": [],
                        }
                    else:
                        # Layer detected but no info extracted - reset current layer
                        current_layer = None
                        current_layer_info = None

                    # Continue to check if this row also contains column headers
                    # (some formats have layer + headers on same row)

                # 4. Check if this is a participant header row
                if is_participant_header_row(row_list):
                    col_map = map_columns(row_list)
                    if DEBUG_PARSING:
                        logger.debug(f"  -> Participant header, col_map: {col_map}")
                    continue

                # 5. Try to extract carrier information
                if current_layer and col_map.get("carrier") is not None:
                    carrier_idx = col_map["carrier"]
                    line_idx = col_map.get("line")

                    if DEBUG_PARSING:
                        carrier_val = (
                            row_list[carrier_idx]
                            if carrier_idx < len(row_list)
                            else None
                        )
                        line_val = (
                            row_list[line_idx]
                            if line_idx and line_idx < len(row_list)
                            else None
                        )
                        logger.debug(
                            f"  -> Reading: carrier_col[{carrier_idx}]='{carrier_val}', line_col[{line_idx}]='{line_val}'"
                        )

                    # Ensure index is within bounds
                    if carrier_idx < len(row_list) and row_list[carrier_idx]:
                        carrier_name = str(row_list[carrier_idx]).strip()

                        # Skip invalid names or headers that might have slipped through
                        skip_names = {
                            "PARTICIPANT",
                            "CARRIER",
                            "TOTAL",
                            "TOTALS",
                            "FIRM",
                            "INSURER",
                            "UNDERWRITER",
                            "GRAND TOTAL",
                            "SUBTOTAL",
                            "N/A",
                            "TBD",
                            # Skip Yes/No values (checkbox columns)
                            "YES",
                            "NO",
                            "Y",
                            "N",
                            "TRUE",
                            "FALSE",
                            # Skip header words
                            "LINE",
                            "PREMIUM",
                            "FEES",
                            "FEE",
                            "PPM",
                            "RATE",
                            "SL TAX",
                            "SURPLUS",
                        }
                        carrier_upper = carrier_name.upper().strip()
                        if not carrier_name or carrier_upper in skip_names:
                            if DEBUG_PARSING:
                                logger.debug(
                                    f"  -> Skipping invalid/skip name: '{carrier_name}'"
                                )
                            continue

                        # Skip if carrier name is too short (likely not a real carrier)
                        # Real carrier names are typically at least 3 characters
                        if len(carrier_name.strip()) < 3:
                            if DEBUG_PARSING:
                                logger.debug(
                                    f"  -> Skipping too-short name: '{carrier_name}'"
                                )
                            continue

                        # Skip if carrier name looks like a number or currency
                        if re.match(r"^[\$\d,.\-\s%]+$", carrier_name.strip()):
                            if DEBUG_PARSING:
                                logger.debug(
                                    f"  -> Skipping numeric value: '{carrier_name}'"
                                )
                            continue

                        # Valid carrier names typically contain letters and often
                        # have words like "Insurance", "Company", "Underwriters", etc.
                        # But we don't require these - just ensure it's not garbage

                        # Extract line (participation amount)
                        line_amount = 0
                        if col_map.get("line") is not None and col_map["line"] < len(
                            row_list
                        ):
                            line_amount = parse_currency(row_list[col_map["line"]])

                        # Extract premium
                        premium = 0
                        if col_map.get("premium") is not None and col_map[
                            "premium"
                        ] < len(row_list):
                            premium = parse_currency(row_list[col_map["premium"]])

                        # If no dedicated premium column, try total column
                        if (
                            premium == 0
                            and col_map.get("total") is not None
                            and col_map["total"] < len(row_list)
                        ):
                            premium = parse_currency(row_list[col_map["total"]])

                        # Extract fees
                        carrier_fee = 0
                        if col_map.get("fees") is not None and col_map["fees"] < len(
                            row_list
                        ):
                            carrier_fee = parse_currency(row_list[col_map["fees"]])

                        # Extract SL tax
                        surplus_fee = 0
                        if col_map.get("sl_tax") is not None and col_map[
                            "sl_tax"
                        ] < len(row_list):
                            surplus_fee = parse_currency(row_list[col_map["sl_tax"]])

                        # Calculate share: Line Amount / Layer Limit
                        # Use the layer's limit directly from current_layer for consistency
                        layer_limit = current_layer.get("limit", 0)
                        share = 0.0
                        if layer_limit > 0 and line_amount > 0:
                            share = line_amount / layer_limit

                        if DEBUG_PARSING:
                            logger.debug(
                                f"  -> Share calc: {line_amount:,.0f} / {layer_limit:,.0f} = {share:.4f} ({share*100:.2f}%)"
                            )

                        # Add carrier to current layer if they have participation
                        if carrier_name and (line_amount > 0 or premium > 0):
                            carrier_data = {
                                "carrier_name": carrier_name,
                                "share": share,
                                "premium": premium,
                                "carrier_fee": carrier_fee,
                                "surplus_fee": surplus_fee,
                                "policy_number": "",
                                "has_multiple_rbes": False,
                                "rbes": [],
                            }
                            current_layer["carriers"].append(carrier_data)

                            if DEBUG_PARSING:
                                logger.debug(
                                    f"  -> Added carrier: {carrier_name}, share={share:.4f} ({share*100:.2f}%), line={line_amount:,.0f}, premium={premium:,.0f}"
                                )

        # Add last layer if exists
        if current_layer and current_layer.get("carriers"):
            all_layers.append(current_layer)
            if DEBUG_PARSING:
                logger.debug(
                    f"\nSaved final layer with {len(current_layer['carriers'])} carriers"
                )

        # Merge duplicate layers (same limit + attachment) that may have been split
        # This can happen if layer headers appear multiple times or across sheets
        merged_layers = {}
        for layer in all_layers:
            key = (layer["limit"], layer["attachment"])
            if key not in merged_layers:
                merged_layers[key] = {
                    "limit": layer["limit"],
                    "attachment": layer["attachment"],
                    "is_primary": layer["is_primary"],
                    "carriers": [],
                }
            # Add carriers - allow same carrier name with different line amounts
            # (e.g., Palomar may appear 3 times with different participations)
            for carrier in layer.get("carriers", []):
                # Check if this exact carrier entry already exists (same name AND same share)
                is_duplicate = False
                for existing in merged_layers[key]["carriers"]:
                    if (
                        existing["carrier_name"] == carrier["carrier_name"]
                        and abs(existing["share"] - carrier["share"]) < 0.0001
                    ):
                        is_duplicate = True
                        break
                if not is_duplicate:
                    merged_layers[key]["carriers"].append(carrier)

        # Convert back to list and sort by attachment point
        all_layers = sorted(merged_layers.values(), key=lambda x: x["attachment"])

        # Validate and log share calculations for each layer
        if DEBUG_PARSING:
            logger.debug(f"\n{'='*50}")
            logger.debug(f"Parse complete: {len(all_layers)} layers found")

        for layer in all_layers:
            carriers = layer.get("carriers", [])
            if carriers:
                total_share = sum(c.get("share", 0) for c in carriers)
                layer_limit = layer.get("limit", 0)

                if DEBUG_PARSING:
                    logger.debug(
                        f"\nLayer ${layer['limit']:,.0f} xs ${layer['attachment']:,.0f}:"
                    )
                    for c in carriers:
                        line_calc = c.get("share", 0) * layer_limit
                        logger.debug(
                            f"  - {c['carrier_name']}: {c['share']*100:.2f}% (line ~${line_calc:,.0f})"
                        )
                    logger.debug(f"  TOTAL: {total_share*100:.2f}%")

                    if abs(total_share - 1.0) > 0.01:
                        logger.warning(
                            f"  ⚠️ Shares don't sum to 100%! Total: {total_share*100:.2f}%"
                        )

        return {
            "filename": filename,
            "layers": all_layers,
            "success": True,
            "error": None,
        }

    except Exception as e:
        if DEBUG_PARSING:
            logger.exception(f"Error parsing Excel file: {e}")
        return {"filename": filename, "layers": [], "success": False, "error": str(e)}

    finally:
        DEBUG_PARSING = original_debug


def merge_excel_programs(excel_data_list: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Merge multiple parsed Excel programs into a single program structure.

    Layers are keyed by (limit, attachment) to prevent duplicates.
    Carriers within the same layer are merged by name.

    Args:
        excel_data_list: List of parsed Excel data dictionaries

    Returns:
        Program structure with merged layers
    """
    layer_dict = {}  # Key: (limit, attachment) -> layer data

    for excel_data in excel_data_list:
        if not excel_data.get("success"):
            continue

        for layer in excel_data.get("layers", []):
            key = (layer["limit"], layer["attachment"])

            if key not in layer_dict:
                layer_dict[key] = {
                    "limit": layer["limit"],
                    "attachment": layer["attachment"],
                    "is_primary": layer["is_primary"],
                    "carriers": [],
                }

            # Add carriers from this layer
            # Allow same carrier to appear multiple times with different line amounts
            for carrier in layer.get("carriers", []):
                # Check if this EXACT entry already exists (same name AND same share)
                is_exact_duplicate = False
                for existing in layer_dict[key]["carriers"]:
                    if (
                        existing["carrier_name"] == carrier["carrier_name"]
                        and abs(existing["share"] - carrier.get("share", 0)) < 0.0001
                    ):
                        # Exact duplicate - just sum premiums/fees
                        existing["premium"] += carrier.get("premium", 0)
                        existing["carrier_fee"] += carrier.get("carrier_fee", 0)
                        existing["surplus_fee"] += carrier.get("surplus_fee", 0)
                        is_exact_duplicate = True
                        break

                if not is_exact_duplicate:
                    # New carrier entry (or same carrier with different share)
                    layer_dict[key]["carriers"].append(carrier.copy())

    # Convert to list and sort by attachment
    layers = sorted(layer_dict.values(), key=lambda x: x["attachment"])

    return {
        "layers": layers,
        "documents_processed": len([d for d in excel_data_list if d.get("success")]),
        "documents_failed": len([d for d in excel_data_list if not d.get("success")]),
    }


# Utility function for testing/debugging
def test_layer_parsing():
    """Test function to verify layer pattern extraction."""
    test_cases = [
        "$75M ex $100M EQ",
        "$100M ex $300M",
        "$500M ex $1BL",
        "$500M ex $1.5BL AR ex",
        "$50M ex $250M",
        "$250M Terrorism",
        "ALL RISKS EX ZURICH LEAD",
        "$1BL Primary",
        "75M ex 100M",
        "$75,000,000 ex $100,000,000",
    ]

    print("Layer Parsing Test Results:")
    print("=" * 60)
    for test in test_cases:
        result = extract_layer_from_text(test)
        if result:
            print(f"✓ '{test}'")
            print(
                f"  -> Limit: ${result['limit']:,.0f}, Attachment: ${result['attachment']:,.0f}, Primary: {result['is_primary']}"
            )
        else:
            print(f"✗ '{test}' -> No match")
        print()


if __name__ == "__main__":
    test_layer_parsing()
