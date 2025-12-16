from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from io import BytesIO


def export_program_to_excel(program, carrier_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Program Structure"

    # Define styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    carrier_fill = PatternFill(
        start_color="EBF1FA", end_color="EBF1FA", fill_type="solid"
    )
    rbe_header_fill = PatternFill(
        start_color="E6EEF9", end_color="E6EEF9", fill_type="solid"
    )
    rbe_fill = PatternFill(start_color="F5F9FF", end_color="F5F9FF", fill_type="solid")
    separator_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(size=16, bold=True)
    carrier_font = Font(bold=True)
    rbe_header_font = Font(italic=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    bottom_border = Border(bottom=Side(style="thin"))
    top_border = Border(top=Side(style="thin"))

    # Title
    ws["A1"] = program["account"]
    ws["A1"].font = title_font
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3

    # Sort layers by attachment
    layers_sorted = sorted(program["layers"], key=lambda x: x.get("attachment", 0))

    for idx, layer in enumerate(layers_sorted):
        limit_val = layer.get("limit", 0)
        attach_val = layer.get("attachment", 0)

        if layer.get("is_primary"):
            layer_title = f"Layer {idx+1}: ${limit_val:,.0f} Primary"
        else:
            layer_title = f"Layer {idx+1}: ${limit_val:,.0f} xs ${attach_val:,.0f}"

        # Layer title row with improved styling
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.border = thin_border

        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = layer_title
        cell.font = header_font
        # Enable text wrapping to prevent cutoff
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.row_dimensions[row].height = 35  # Increased header row height
        row += 1

        # Column headers with improved styling
        headers = ["Carrier", "Share %", "Premium ($)", "Policy #", "Total Fees ($)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = subheader_fill
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        # Process carriers
        for carrier_idx, carrier in enumerate(layer.get("carriers", [])):
            # Add clear separator before each carrier (except the first one)
            if carrier_idx > 0:
                # Add a visible separator row
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = separator_fill
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
                ws.row_dimensions[row].height = 5  # Thin separator row
                row += 1

            carrier_name = carrier.get("carrier_name", "Unknown")
            carrier_share = carrier.get("share", 0)
            premium = carrier.get("premium", 0)
            policy_num = carrier.get("policy_number", "")
            carrier_fee = carrier.get("carrier_fee", 0)
            surplus_fee = carrier.get("surplus_fee", 0)
            total_fees = carrier_fee + surplus_fee

            # Handle Multiple RBEs policy number display
            if (
                carrier.get("has_multiple_rbes", False)
                and not carrier.get("single_policy_number", False)
                and carrier.get("rbes", [])
            ):
                policy_num = "Multiple"

            # Carrier row with highlighting
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = carrier_fill
                cell.border = thin_border

            ws.cell(row=row, column=1, value=carrier_name).font = carrier_font
            ws.cell(row=row, column=2, value=f"{carrier_share*100:.1f}%")
            ws.cell(row=row, column=3, value=f"${premium:,.0f}")
            ws.cell(row=row, column=4, value=policy_num)
            ws.cell(row=row, column=5, value=f"${total_fees:,.2f}")
            ws.row_dimensions[row].height = 20  # Taller carrier rows
            row += 1

            if carrier.get("has_multiple_rbes", False):
                # RBE breakdown header with distinct styling
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin", color="D0D0D0"),
                        bottom=Side(style="thin", color="D0D0D0"),
                    )

                ws.merge_cells(f"A{row}:E{row}")
                cell = ws[f"A{row}"]
                cell.value = f"RBE breakdown of carrier's {carrier_share*100:.1f}% layer participation:"
                cell.font = rbe_header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = rbe_header_fill
                row += 1

                # RBE column headers
                rbe_headers = [
                    "RBE",
                    "RBE Share %",
                    "Layer Share %",
                    "Premium ($)",
                    "Policy #",
                ]
                for col_idx, header in enumerate(rbe_headers, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.fill = rbe_header_fill
                    cell.font = Font(bold=True, size=9)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                row += 1

                # RBE rows with alternating subtle colors
                for rbe_idx, rbe in enumerate(carrier.get("rbes", [])):
                    rbe_name = rbe.get("rbe", "")
                    rbe_share = rbe.get("share", 0) * 100
                    layer_share = rbe.get("share", 0) * carrier_share * 100
                    premium = rbe.get("premium", 0)

                    # Use carrier policy number if single policy is enabled
                    rbe_policy_num = rbe.get("policy_number", "")
                    if carrier.get("single_policy_number", False):
                        rbe_policy_num = carrier.get("policy_number", "")

                    # Alternate subtle background for RBE rows
                    for col in range(1, 6):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = rbe_fill
                        cell.border = thin_border
                        # Indent the first column for RBEs
                        if col == 1:
                            cell.alignment = Alignment(indent=2)

                    ws.cell(row=row, column=1, value=rbe_name)
                    ws.cell(row=row, column=2, value=f"{rbe_share:.1f}%")
                    ws.cell(row=row, column=3, value=f"{layer_share:.2f}%")
                    ws.cell(row=row, column=4, value=f"${premium:,.0f}")
                    ws.cell(row=row, column=5, value=rbe_policy_num)
                    row += 1

                # Add a closing border for the RBE section
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = Border(bottom=Side(style="thin", color="D0D0D0"))
                    cell.fill = rbe_header_fill
                ws.row_dimensions[row].height = 5  # Thin closing line
                row += 1

        # Add space between layers
        row += 1
        for col in range(1, 6):
            # Add a subtle separator between layers
            cell = ws.cell(row=row, column=col)
            cell.border = Border(bottom=Side(style="dashed", color="D0D0D0"))
        row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 18

        # Add a light grid to the entire sheet
        for row_idx in range(1, row):
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.border:
                    cell.border = Border(
                        left=Side(style="thin", color="E0E0E0"),
                        right=Side(style="thin", color="E0E0E0"),
                        top=Side(style="thin", color="E0E0E0"),
                        bottom=Side(style="thin", color="E0E0E0"),
                    )

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
